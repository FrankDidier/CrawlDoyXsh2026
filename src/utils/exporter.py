"""
Export functionality for parsed results.
"""

import csv
import json
from typing import List
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..parsers.base import ParseResult


class Exporter:
    """Export parsed results to various formats"""
    
    @staticmethod
    def to_csv(results: List[ParseResult], filepath: str) -> bool:
        """Export results to CSV file"""
        try:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                if not results:
                    return False
                
                fieldnames = list(results[0].to_dict().keys())
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for result in results:
                    writer.writerow(result.to_dict())
            
            return True
        except Exception as e:
            print(f"CSV export error: {e}")
            return False
    
    @staticmethod
    def to_excel(results: List[ParseResult], filepath: str) -> bool:
        """Export results to Excel file"""
        if not HAS_OPENPYXL:
            print("openpyxl not installed. Please install it: pip install openpyxl")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "提取结果"
            
            if not results:
                return False
            
            # Headers
            headers = list(results[0].to_dict().keys())
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data
            for row_idx, result in enumerate(results, 2):
                data = result.to_dict()
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=data.get(header, ""))
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Excel export error: {e}")
            return False
    
    @staticmethod
    def to_json(results: List[ParseResult], filepath: str) -> bool:
        """Export results to JSON file"""
        try:
            data = [result.to_dict() for result in results]
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"JSON export error: {e}")
            return False
    
    @staticmethod
    def generate_filename(prefix: str = "export", extension: str = "xlsx") -> str:
        """Generate a filename with timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"
    
    @staticmethod
    def to_excel_from_dicts(data: list, filepath: str) -> bool:
        """Export list of dictionaries to Excel"""
        if not HAS_OPENPYXL:
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "抓取结果"
            
            if not data:
                return False
            
            # Headers
            headers = list(data[0].keys())
            
            # Style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                    cell.border = thin_border
            
            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Excel export error: {e}")
            return False
    
    @staticmethod
    def to_csv_from_dicts(data: list, filepath: str) -> bool:
        """Export list of dictionaries to CSV"""
        try:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                if not data:
                    return False
                
                fieldnames = list(data[0].keys())
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for row in data:
                    writer.writerow(row)
            
            return True
        except Exception as e:
            print(f"CSV export error: {e}")
            return False
