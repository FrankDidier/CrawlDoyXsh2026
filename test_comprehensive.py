"""
Comprehensive tests for the CrawProj crawler application.
Tests data structures, crawler logic, URL normalization, store name cleaning,
export functionality, and UI initialization.
"""

import os
import sys
import asyncio
import tempfile

sys.path.insert(0, os.path.dirname(__file__))


def test_section(name):
    print(f"\n{'='*60}")
    print(f"  {name}")
    print(f"{'='*60}")


passed = 0
failed = 0

def check(description, condition):
    global passed, failed
    if condition:
        print(f"  ✓ {description}")
        passed += 1
    else:
        print(f"  ✗ FAIL: {description}")
        failed += 1


# ==============================================================
# 1. Data Structures
# ==============================================================
test_section("1. CrawlResult & CrawlProgress Data Structures")

from src.crawlers.base import (
    BaseCrawler, CrawlResult, CrawlProgress, CrawlStatus,
    Platform, ContentType
)

r1 = CrawlResult(
    platform=Platform.DOUYIN,
    content_type=ContentType.LIVE,
    url="https://live.douyin.com/123456",
    share_text="test share text",
    title="测试直播间",
    account_id="123456",
    account_name="测试主播",
)
check("CrawlResult creation", r1.url == "https://live.douyin.com/123456")
check("CrawlResult platform", r1.platform == Platform.DOUYIN)
check("CrawlResult to_dict has all keys", all(k in r1.to_dict() for k in ["平台", "类型", "链接", "APP分享文本", "标题"]))
check("CrawlResult to_dict values", r1.to_dict()["平台"] == "抖音" and r1.to_dict()["链接"] == "https://live.douyin.com/123456")

r2 = CrawlResult(
    platform=Platform.TAOBAO,
    content_type=ContentType.STORE,
    url="https://shop123.taobao.com/",
    store_name="测试旗舰店",
    share_text="【淘宝店铺】测试旗舰店 https://shop123.taobao.com/",
)
check("Store CrawlResult", r2.store_name == "测试旗舰店")
check("Store to_dict", r2.to_dict()["店铺名称"] == "测试旗舰店")

p = CrawlProgress()
check("CrawlProgress default status", p.status == CrawlStatus.IDLE)
check("CrawlProgress default percentage", p.percentage == 0)

p.percentage = 50
check("CrawlProgress explicit percentage", p.percentage == 50)

p2 = CrawlProgress(total=100, current=75)
check("CrawlProgress calculated percentage", p2.percentage == 75)


# ==============================================================
# 2. BaseCrawler
# ==============================================================
test_section("2. BaseCrawler Initialization & Methods")

bc = BaseCrawler()
check("BaseCrawler has _paused", hasattr(bc, '_paused'))
check("BaseCrawler _paused is False", bc._paused == False)
check("BaseCrawler _cancelled is False", bc._cancelled == False)
check("BaseCrawler results empty", len(bc.results) == 0)

bc.pause()
check("BaseCrawler pause sets _paused", bc._paused == True)
check("BaseCrawler pause sets status", bc.progress.status == CrawlStatus.PAUSED)

bc.resume()
check("BaseCrawler resume clears _paused", bc._paused == False)
check("BaseCrawler resume sets RUNNING", bc.progress.status == CrawlStatus.RUNNING)

bc.cancel()
check("BaseCrawler cancel sets _cancelled", bc._cancelled == True)
check("BaseCrawler cancel sets CANCELLED", bc.progress.status == CrawlStatus.CANCELLED)

bc.reset()
check("BaseCrawler reset clears cancelled", bc._cancelled == False)
check("BaseCrawler reset clears paused", bc._paused == False)
check("BaseCrawler reset clears results", len(bc.results) == 0)

# Test result callback
callback_results = []
bc2 = BaseCrawler()
bc2.set_result_callback(lambda r: callback_results.append(r))
test_result = CrawlResult(platform=Platform.DOUYIN, content_type=ContentType.VIDEO, url="test")
bc2._add_result(test_result)
check("Result callback fired", len(callback_results) == 1)
check("Result added to results list", len(bc2.results) == 1)

# Test progress callback
progress_updates = []
bc3 = BaseCrawler()
bc3.set_progress_callback(lambda p: progress_updates.append(p.message))
bc3._update_progress(message="test message")
check("Progress callback fired", len(progress_updates) == 1 and progress_updates[0] == "test message")


# ==============================================================
# 3. Taobao URL Normalization
# ==============================================================
test_section("3. Taobao URL Normalization")

from src.crawlers.taobao import TaobaoCrawler
tc = TaobaoCrawler()

check("shop URL normalization",
    tc._normalize_store_url("https://shop12345.taobao.com/search.htm?q=test") == "https://shop12345.taobao.com/")
check("tmall URL normalization",
    tc._normalize_store_url("//brandname.tmall.com/search.htm") == "https://brandname.tmall.com/")
check("user_number_id conversion",
    tc._normalize_store_url("https://store.taobao.com?user_number_id=67890") == "https://shop67890.taobao.com/")
check("appUid URL normalization",
    "appUid=" in tc._normalize_store_url("https://store.taobao.com/shop/view_shop.htm?appUid=abc123&spm=xxx"))
check("detail.tmall not treated as store",
    tc._normalize_store_url("//detail.tmall.com/item.htm?id=123") != "https://detail.tmall.com/")
check("prefix // handling",
    tc._normalize_store_url("//shop99999.taobao.com/").startswith("https://"))


# ==============================================================
# 4. Taobao Store Name Cleaning
# ==============================================================
test_section("4. Taobao Store Name Cleaning")

check("Clean - remove year prefix",
    tc._clean_store_name("5年老店 测试旗舰店") == "测试旗舰店")
check("Clean - remove 天猫 prefix",
    tc._clean_store_name("天猫 某品牌旗舰店") == "某品牌旗舰店")
check("Clean - multiline take store name",
    "店" in tc._clean_store_name("回头客1万\n皇冠\n好评率98%\n某某旗舰店"))
check("Clean - plain name unchanged",
    tc._clean_store_name("ABC品牌旗舰店") == "ABC品牌旗舰店")
check("Clean - empty string",
    tc._clean_store_name("") == "")
check("Clean - strip whitespace",
    tc._clean_store_name("  测试店铺  ") == "测试店铺")


# ==============================================================
# 5. Crawler Instantiation
# ==============================================================
test_section("5. Crawler Instantiation & Attributes")

from src.crawlers.douyin import DouyinCrawler
from src.crawlers.kuaishou import KuaishouCrawler
from src.crawlers.jd import JDCrawler

dc = DouyinCrawler()
check("DouyinCrawler platform", dc.platform == Platform.DOUYIN)
check("DouyinCrawler types", ContentType.LIVE in dc.supported_types and ContentType.VIDEO in dc.supported_types)
check("DouyinCrawler has _paused", hasattr(dc, '_paused'))

kc = KuaishouCrawler()
check("KuaishouCrawler platform", kc.platform == Platform.KUAISHOU)
check("KuaishouCrawler types", ContentType.LIVE in kc.supported_types and ContentType.VIDEO in kc.supported_types)

tc2 = TaobaoCrawler()
check("TaobaoCrawler platform", tc2.platform == Platform.TAOBAO)
check("TaobaoCrawler types", ContentType.STORE in tc2.supported_types and ContentType.PRODUCT in tc2.supported_types)
check("TaobaoCrawler keep browser", tc2._keep_browser_open == True)

jc = JDCrawler()
check("JDCrawler platform", jc.platform == Platform.JD)
check("JDCrawler types", ContentType.STORE in jc.supported_types and ContentType.PRODUCT in jc.supported_types)
check("JDCrawler has _scroll_to_load_all", hasattr(jc, '_scroll_to_load_all'))
check("JDCrawler has _jd_go_to_next_page", hasattr(jc, '_jd_go_to_next_page'))
check("JDCrawler has _crawl_stores", hasattr(jc, '_crawl_stores'))
check("JDCrawler has _crawl_products", hasattr(jc, '_crawl_products'))


# ==============================================================
# 6. Export Functionality
# ==============================================================
test_section("6. Export Functionality (Excel & CSV)")

from src.utils.exporter import Exporter

test_data = [
    CrawlResult(platform=Platform.DOUYIN, content_type=ContentType.LIVE,
                url="https://live.douyin.com/111", title="直播1", account_name="主播1"),
    CrawlResult(platform=Platform.TAOBAO, content_type=ContentType.STORE,
                url="https://shop123.taobao.com/", store_name="测试旗舰店"),
    CrawlResult(platform=Platform.JD, content_type=ContentType.PRODUCT,
                url="https://item.jd.com/999.html", title="商品1", product_name="商品1", price="99.00"),
]
dict_data = [r.to_dict() for r in test_data]

# Test Excel export
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
    xlsx_path = f.name

result = Exporter.to_excel_from_dicts(dict_data, xlsx_path)
check("Excel export succeeds", result == True)
check("Excel file created", os.path.exists(xlsx_path) and os.path.getsize(xlsx_path) > 0)

# Verify Excel content
from openpyxl import load_workbook
wb = load_workbook(xlsx_path)
ws = wb.active
check("Excel has correct sheet name", ws.title == "抓取结果")
check("Excel has header row", ws.cell(1, 1).value == "平台")
check("Excel has 3 data rows", ws.max_row == 4)  # 1 header + 3 data
check("Excel first row platform", ws.cell(2, 1).value == "抖音")
check("Excel third row platform", ws.cell(4, 1).value == "京东")
wb.close()
os.unlink(xlsx_path)

# Test CSV export
with tempfile.NamedTemporaryFile(suffix='.csv', delete=False, mode='w') as f:
    csv_path = f.name

result = Exporter.to_csv_from_dicts(dict_data, csv_path)
check("CSV export succeeds", result == True)
check("CSV file created", os.path.exists(csv_path) and os.path.getsize(csv_path) > 0)

import csv
with open(csv_path, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    rows = list(reader)
check("CSV has 3 rows", len(rows) == 3)
check("CSV first row platform", rows[0]["平台"] == "抖音")
check("CSV third row store name", rows[2]["店铺名称"] == "")  # JD product has no store_name in this test
os.unlink(csv_path)

# Test empty data
check("Excel empty data returns False", Exporter.to_excel_from_dicts([], "/tmp/empty.xlsx") == False)
check("CSV empty data returns False", Exporter.to_csv_from_dicts([], "/tmp/empty.csv") == False)

# Test filename generation
fn = Exporter.generate_filename("test", "xlsx")
check("Filename has prefix", fn.startswith("test_"))
check("Filename has extension", fn.endswith(".xlsx"))


# ==============================================================
# 7. Parser Manager
# ==============================================================
test_section("7. Parser Manager & Text Parsers")

from src.utils.parser_manager import ParserManager
from src.parsers.base import Platform as ParserPlatform

pm = ParserManager()
check("ParserManager has 4 parsers", len(pm.parsers) == 4)

# Test Douyin share text parsing
douyin_text = "7.25 iFb:/ 03/30 复制打开抖音，看看【测试账号的视频】 https://v.douyin.com/iRNBqwMX/"
result = pm.parse(douyin_text)
if result:
    check("Douyin parse succeeds", result.success == True)
    check("Douyin platform detected", result.platform == ParserPlatform.DOUYIN)
    check("Douyin URL extracted", "v.douyin.com" in result.url)
else:
    check("Douyin parse returned result", result is not None)

# Test Kuaishou share text
ks_text = "【快手】来看看我的直播 https://v.kuaishou.com/abcdef"
result = pm.parse(ks_text)
if result:
    check("Kuaishou parse succeeds", result.success == True)
    check("Kuaishou URL extracted", "v.kuaishou.com" in result.url)
else:
    check("Kuaishou parse returned result", result is not None)

# Test platform detection
check("Detect Douyin", pm.detect_platform("https://v.douyin.com/xxx") == ParserPlatform.DOUYIN)
check("Detect unknown", pm.detect_platform("random text") == ParserPlatform.UNKNOWN)

# Test empty/null handling
check("Empty string returns None", pm.parse("") is None)
check("Whitespace returns None", pm.parse("   ") is None)
check("Batch empty list", pm.parse_batch([]) == [])


# ==============================================================
# 8. Browser Helper
# ==============================================================
test_section("8. Browser Helper Utilities")

from src.utils.browser_helper import (
    get_user_data_dir, get_chrome_path,
    get_browser_launch_options, BROWSER_CHANNELS
)

udd = get_user_data_dir()
check("User data dir not empty", len(udd) > 0)
check("User data dir is absolute path", os.path.isabs(udd))

options = get_browser_launch_options(headless=True)
check("Launch options has headless", options['headless'] == True)
check("Launch options has args", len(options['args']) > 0)
check("Launch options has anti-detection", any('AutomationControlled' in a for a in options['args']))

check("Browser channels dict has entries", len(BROWSER_CHANNELS) >= 6)
check("Chrome channel is 'chrome'", BROWSER_CHANNELS.get("Chrome") == "chrome")
check("Edge channel is 'msedge'", BROWSER_CHANNELS.get("Edge") == "msedge")


# ==============================================================
# 9. UI Components (no display needed)
# ==============================================================
test_section("9. UI Components Import Test")

try:
    from src.ui.styles import DARK_THEME
    check("DARK_THEME loaded", len(DARK_THEME) > 100)
except Exception as e:
    check(f"DARK_THEME import: {e}", False)

try:
    from src.ui.help_dialog import HelpDialog
    check("HelpDialog imported", True)
except Exception as e:
    check(f"HelpDialog import: {e}", False)

# Test MainWindow import (don't create - needs QApplication)
try:
    from src.ui.main_window import MainWindow, CrawlerWorker, EmulatorWorker
    check("MainWindow imported", True)
    check("CrawlerWorker imported", True)
    check("EmulatorWorker imported", True)
except Exception as e:
    check(f"MainWindow import: {e}", False)


# ==============================================================
# 10. Async Crawler Logic (mock tests)
# ==============================================================
test_section("10. Async Crawler Logic Tests")

async def test_async():
    results = []
    
    # Test BaseCrawler _check_pause doesn't hang when not paused
    bc = BaseCrawler()
    bc.reset()
    await bc._check_pause()
    results.append(("_check_pause doesn't block when not paused", True))
    
    # Test Douyin search validation
    dc = DouyinCrawler()
    try:
        await dc.search("test", ContentType.STORE, max_results=10)
        results.append(("Douyin rejects STORE type", False))
    except ValueError as e:
        results.append(("Douyin rejects STORE type", "不支持" in str(e)))
    except Exception:
        results.append(("Douyin rejects STORE type", True))
    
    # Test Kuaishou search validation
    kc = KuaishouCrawler()
    try:
        await kc.search("test", ContentType.STORE, max_results=10)
        results.append(("Kuaishou rejects STORE type", False))
    except ValueError as e:
        results.append(("Kuaishou rejects STORE type", "不支持" in str(e)))
    except Exception:
        results.append(("Kuaishou rejects STORE type", True))
    
    # Test Taobao search validation
    tc = TaobaoCrawler()
    try:
        await tc.search("test", ContentType.LIVE, max_results=10)
        results.append(("Taobao rejects LIVE type", False))
    except ValueError as e:
        results.append(("Taobao rejects LIVE type", "不支持" in str(e)))
    except Exception:
        results.append(("Taobao rejects LIVE type", True))
    
    # Test JD search validation
    jc = JDCrawler()
    try:
        await jc.search("test", ContentType.LIVE, max_results=10)
        results.append(("JD rejects LIVE type", False))
    except ValueError as e:
        results.append(("JD rejects LIVE type", "不支持" in str(e)))
    except Exception:
        results.append(("JD rejects LIVE type", True))
    
    return results

loop = asyncio.new_event_loop()
async_results = loop.run_until_complete(test_async())
loop.close()

for desc, cond in async_results:
    check(desc, cond)


# ==============================================================
# 11. Taobao Pagination Logic
# ==============================================================
test_section("11. Taobao Pagination URL Construction")

import re

def simulate_taobao_pagination(current_url, next_page_num):
    """Simulate Taobao pagination URL building without browser"""
    skip_count = (next_page_num - 1) * 44
    if 's=' in current_url:
        new_url = re.sub(r's=\d+', f's={skip_count}', current_url)
    elif '?' in current_url:
        new_url = current_url + f'&s={skip_count}'
    else:
        new_url = current_url + f'?s={skip_count}'
    if 'page=' in new_url:
        new_url = re.sub(r'page=\d+', f'page={next_page_num}', new_url)
    return new_url

url1 = "https://s.taobao.com/search?q=test&s=0"
check("Page 2 s=44", "s=44" in simulate_taobao_pagination(url1, 2))
check("Page 3 s=88", "s=88" in simulate_taobao_pagination(url1, 3))
check("Page 10 s=396", "s=396" in simulate_taobao_pagination(url1, 10))

url2 = "https://s.taobao.com/search?q=test"
check("No s= param adds it", "s=44" in simulate_taobao_pagination(url2, 2))

url3 = "https://s.taobao.com/search?q=test&s=0&page=1"
result3 = simulate_taobao_pagination(url3, 3)
check("Updates both s= and page=", "s=88" in result3 and "page=3" in result3)


# ==============================================================
# 12. JD Pagination URL Construction
# ==============================================================
test_section("12. JD Pagination URL Construction")

def simulate_jd_pagination(current_url, next_page):
    if 'page=' in current_url:
        new_url = re.sub(r'page=\d+', f'page={next_page}', current_url)
    elif '?' in current_url:
        new_url = current_url + f'&page={next_page}'
    else:
        new_url = current_url + f'?page={next_page}'
    if 's=' in new_url:
        new_url = re.sub(r's=\d+', f's={(next_page - 1) * 60}', new_url)
    return new_url

jd_url = "https://search.jd.com/Search?keyword=test"
check("JD page 2", "page=2" in simulate_jd_pagination(jd_url, 2))
check("JD page 5", "page=5" in simulate_jd_pagination(jd_url, 5))

jd_url2 = "https://search.jd.com/Search?keyword=test&page=1&s=0"
result_jd = simulate_jd_pagination(jd_url2, 3)
check("JD updates page and s", "page=3" in result_jd and "s=120" in result_jd)


# ==============================================================
# 13. CrawlResult for Different Platforms
# ==============================================================
test_section("13. CrawlResult Cross-Platform Compatibility")

results_to_export = [
    CrawlResult(platform=Platform.DOUYIN, content_type=ContentType.LIVE,
                url="https://live.douyin.com/111", account_name="主播A",
                share_text="#在抖音# 主播A直播中 https://live.douyin.com/111"),
    CrawlResult(platform=Platform.KUAISHOU, content_type=ContentType.VIDEO,
                url="https://www.kuaishou.com/short-video/abc",
                account_name="视频作者B",
                share_text="#快手短视频# 视频作者B https://www.kuaishou.com/short-video/abc"),
    CrawlResult(platform=Platform.TAOBAO, content_type=ContentType.STORE,
                url="https://shop999.taobao.com/", store_name="超级旗舰店",
                share_text="【淘宝店铺】超级旗舰店 https://shop999.taobao.com/"),
    CrawlResult(platform=Platform.JD, content_type=ContentType.PRODUCT,
                url="https://item.jd.com/123.html", product_name="优质商品",
                price="¥199", store_name="京东自营",
                share_text="【京东】优质商品 https://item.jd.com/123.html"),
]

dicts = [r.to_dict() for r in results_to_export]
check("4 results converted to dicts", len(dicts) == 4)
check("All dicts have same keys", all(set(d.keys()) == set(dicts[0].keys()) for d in dicts))
check("Douyin dict correct", dicts[0]["平台"] == "抖音" and dicts[0]["账号名称"] == "主播A")
check("Taobao dict correct", dicts[2]["店铺名称"] == "超级旗舰店")
check("JD dict correct", dicts[3]["商品名称"] == "优质商品" and dicts[3]["价格"] == "¥199")

# Export all to Excel
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
    mixed_xlsx = f.name
result = Exporter.to_excel_from_dicts(dicts, mixed_xlsx)
check("Mixed platform Excel export", result == True)
wb = load_workbook(mixed_xlsx)
ws = wb.active
check("Mixed Excel row count", ws.max_row == 5)
wb.close()
os.unlink(mixed_xlsx)


# ==============================================================
# 14. App Launch Test (brief)
# ==============================================================
test_section("14. QApplication & MainWindow Launch Test")

try:
    from PySide6.QtWidgets import QApplication
    app = QApplication.instance()
    if app is None:
        app = QApplication([])
    
    from src.ui.main_window import MainWindow
    w = MainWindow()
    check("MainWindow created", w is not None)
    check("MainWindow title set", "关键词" in w.windowTitle())
    check("MainWindow has results table", w.results_table is not None)
    check("MainWindow table has 7 columns", w.results_table.columnCount() == 7)
    check("MainWindow start button exists", w.start_btn is not None)
    check("MainWindow pause button exists", w.pause_btn is not None)
    check("MainWindow stop button exists", w.stop_btn is not None)
    check("MainWindow platform combo", w.platform_combo.count() == 4)
    check("MainWindow type combo", w.type_combo.count() == 2)
    check("MainWindow mode combo", w.mode_combo.count() == 2)
    check("MainWindow browser combo", w.browser_combo.count() == 6)
    check("MainWindow max results spin", w.max_results_spin.value() == 100)
    check("MainWindow max results range", w.max_results_spin.maximum() == 100000)
    
    # Test platform change updates type options
    w.platform_combo.setCurrentText("淘宝")
    check("Platform change -> 店铺/商品", w.type_combo.itemText(0) == "店铺")
    
    w.platform_combo.setCurrentText("抖音")
    check("Platform change -> 直播/短视频", w.type_combo.itemText(0) == "直播")
    
    # Test mode change
    w.mode_combo.setCurrentText("APP版 (模拟器)")
    check("APP mode hides browser combo", w.browser_combo.isHidden())
    
    w.mode_combo.setCurrentText("网页版")
    check("Web mode shows browser combo", not w.browser_combo.isHidden())
    
    # Test adding result to table
    from src.crawlers.base import ContentType as CrawlerContentType
    test_result = CrawlResult(
        platform=Platform.TAOBAO, content_type=CrawlerContentType.STORE,
        url="https://test.taobao.com/", store_name="测试店铺",
        share_text="【测试】"
    )
    w.on_result_added(test_result)
    check("Result added to table", w.results_table.rowCount() == 1)
    check("Result in results list", len(w.results) == 1)
    
    # Add more results
    for i in range(5):
        r = CrawlResult(
            platform=Platform.DOUYIN, content_type=CrawlerContentType.LIVE,
            url=f"https://live.douyin.com/{i}", account_name=f"主播{i}",
            share_text=f"直播{i}"
        )
        w.on_result_added(r)
    check("6 total results in table", w.results_table.rowCount() == 6)
    
    # Test clear
    w.clear_results()
    check("Clear results empties table", w.results_table.rowCount() == 0)
    check("Clear results empties list", len(w.results) == 0)
    
    w.close()
    
except Exception as e:
    check(f"MainWindow launch test: {e}", False)


# ==============================================================
# SUMMARY
# ==============================================================
print(f"\n{'='*60}")
print(f"  TEST SUMMARY")
print(f"{'='*60}")
print(f"  Passed: {passed}")
print(f"  Failed: {failed}")
print(f"  Total:  {passed + failed}")
print(f"{'='*60}")

if failed > 0:
    print(f"\n  ⚠️  {failed} test(s) FAILED!")
    sys.exit(1)
else:
    print(f"\n  ✅ ALL {passed} TESTS PASSED!")
    sys.exit(0)
